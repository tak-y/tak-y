import sys, os, os.path
from flask import *
from werkzeug.utils import secure_filename
import pandas as pd
import folium
import shutil
from folium import plugins
from livereload import Server
import googlemaps
import branca
import math
import openpyxl
import geocoder

app = Flask(__name__)
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'
allowed_ext = {'csv', 'xlsx', 'xls'}

UPLOAD_FOLDER = 'C:/Users/tkgle/OneDrive/デスクトップ/blog_map/upload'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
def allowed_file(filename):
    return '.' in request.files['file'] and filename.rsplit(',',1).lower() in allowed_ext
@app.route('/')
def index():
    return render_template('index.html')
@app.route('/visual', methods=['GET','POST'])
def visual():
    if request.method == 'POST':
        googleapikey = 'AIzaSyC8rTI8Yv1LrEDnRJ109DfpUsdaQBXAfhE'
        gmaps = googlemaps.Client(key=googleapikey)
        file = request.files['file']
        print(file)
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        if filename.split('.')[1] in allowed_ext:

            flash('ファイルがアップされました')
            wb=openpyxl.load_workbook(file)
            ws = wb["Sheet1"]
            popup = str(request.form['name'])
            link = str(request.form['data'])
            #if bool(circle) == True:
                #data = str(request.form['data'])
                #dim = str(request.form['dim'])
                #div = int(request.form['div'])
                #m = folium.Map(location=[df["緯度"][1],df["経度"][1]],zoom_start=13)
                #for i in range(1,len(df)-1):
                    #print(df[data][i])
                    #folium.Circle(location=[df["緯度"][i],df["経度"][i]],radius=int(df[data][i])/div,color="blue",fill_color=False,popup=str(df[data][i])+dim,tooltip=str(df[name][i])).add_to(m)


                #m.save('C:/Users/tkgle/OneDrive/デスクトップ/data_visualization/static/data_r.html')
                #os.remove(os.path.join('upload', filename))
                #return render_template('visual.html')
            num = math.floor((ws.max_row-1)/2)+1
            print('saidai'+str(ws.max_row))
            m = folium.Map(location=[geocoder.osm(ws.cell(row=2,column=1).value,timeout=1.0).latlng[0],geocoder.osm(ws.cell(row=2,column=1).value,timeout=1.0).latlng[1]],zoom_start=13) 
            for row in ws.iter_rows(min_row=2,max_col=2,max_row=ws.max_row):
                for cell in row:
                    if cell.row<ws.max_row:
                        print(geocoder.osm(ws.cell(row=cell.row,column=1).value,timeout=5.0).latlng)
                        url=str(ws.cell(row=cell.row,column=2).value)
                        print(url)
                        pop=str(ws.cell(row=cell.row,column=1).value)

                        html='<a href ="{url}" target="_blank" rel="noopener noreferrer">{pop}</a>'.format(url=url,pop=pop)
                        # ポップアップの設定・設置
                        iframe = branca.element.IFrame(html=html, width=300, height=100)
                        popup = folium.Popup(iframe, max_width=300,show=True)
                            
                        folium.Marker(location=[geocoder.osm(ws.cell(row=cell.row,column=1).value,timeout=1.0).latlng[0],geocoder.osm(ws.cell(row=cell.row,column=1).value,timeout=1.0).latlng[1]],color="blue",popup=popup).add_to(m)
                        m.save('static/data_r.html')
                    else:
                        break

                    #elif bool(density) == True:
                    #m = folium.Map(location=[df["緯度"][1],df["経度"][1]],zoom_start=13)
                    #cluster = plugins.MarkerCluster().add_to(m)
                        #for i in range(1,len(df)-1):
                           #folium.CircleMarker(location=[df["緯度"][i],df["経度"][i]],color="blue",fill_color=False,tooltip=str(df[name][i])).add_to(cluster)
            return render_template('visual.html')

        else:
            print(error)

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 8080)))





